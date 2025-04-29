"""
    ╔════════════════════════════════════════════════════════════╗
    ║                Модуль utils_file_processor.py              ║
    ╚════════════════════════════════════════════════════════════╝

    Описание:
        Модуль предоставляет функционал для обработки Excel-файлов и поиска товаров:
        • Асинхронная обработка Excel-файлов
        • Поиск товаров с использованием нечеткого сравнения
        • Кэширование результатов для оптимизации производительности
        • Форматирование выходных Excel-файлов
        • Валидация входных данных
        • Отслеживание прогресса обработки
"""

import asyncio
import pandas as pd

from openpyxl        import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
from typing          import Optional, List, Tuple, Callable, Awaitable, Dict
from thefuzz         import fuzz

from src.managers    import EmbeddingManager, DataManager
from src.utils       import preprocessor
from src.utils       import logger

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Регистрируем шрифты с поддержкой кириллицы
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))  # Жирный шрифт

class AsyncCache:
    """
        Простой кэш для асинхронных функций
    """
    def __init__(self, maxsize: int = 1000):
        self.cache: Dict[str, any] = {}
        self.maxsize = maxsize

    def get_key(self, *args, **kwargs) -> str:
        """
            Создает ключ из аргументов
        """
        return f"{args}:{kwargs}"

    async def get_or_create(self, key: str, create_func: Callable) -> any:
        """
            Получает значение из кэша или создает новое
        """
        if key not in self.cache:
            if len(self.cache) >= self.maxsize:
                self.cache.pop(next(iter(self.cache)))
            self.cache[key] = await create_func()
        return self.cache[key]




class ExcelProcessor:
    def __init__(self):
        self.required_columns = {
            'name':     ['наименование', 'название', 'имя', 'name', 'title'],
            'quantity': ['количество', 'кол-во', 'quantity', 'count']
        }
        self.data_manager      = DataManager("data/excel/price-list.xlsx")
        self.embedding_manager = EmbeddingManager(self.data_manager)
        self._progress_callback: Optional[Callable[[float], Awaitable[None]]] = None
        
        self.text_cache       = AsyncCache(maxsize=1000)
        self.similarity_cache = AsyncCache(maxsize=1000)
        
        self.product_types = {
            'лаборатория': ['цифровая лаборатория', 'лабораторное оборудование'],
            'комплект':    ['комплект', 'набор', 'сет'],
            'пособие':     ['пособие', 'материалы', 'комплекс', 'плакаты', 'наглядные']
        }

    def set_progress_callback(self, callback: Callable[[float], Awaitable[None]]):
        """
            Установка callback-функции для отслеживания прогресса
        """
        self._progress_callback = callback

    async def _update_progress(self, progress: float):
        """
            Обновление прогресса через callback
        """
        if self._progress_callback:
            await self._progress_callback(progress)

    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """
            Find column name in DataFrame that matches any of the possible names
        """
        for col in df.columns:
            if any(name.lower() in col.lower() for name in possible_names):
                return col
        return None

    async def preprocess_text(self, text: str) -> str:
        """
            Предварительная обработка текста с использованием TextPreprocessor и кэшированием
        """
        async def _process():
            try:
                return await preprocessor.preprocess(
                    text,
                    remove_stopwords=False,
                    filter_punctuation=True
                )
            except Exception as e:
                logger.error(f"Ошибка при предобработке текста: {e}")
                return ' '.join(text.lower().split())
        return await self.text_cache.get_or_create(text, _process)

    def get_product_type(self, text: str) -> str:
        """
            Определение типа продукта из текста
        """
        text = text.lower()
        for type_name, keywords in self.product_types.items():
            if any(keyword in text for keyword in keywords):
                return type_name
        return "other"

    async def calculate_similarity(self, text1: str, text2: str) -> float:
        """
            Вычисление сходства между двумя текстами с кэшированием
        """
        cache_key = self.similarity_cache.get_key(text1, text2)
        
        async def _calculate():
            try:
                processed_text1 = await self.preprocess_text(text1)
                processed_text2 = await self.preprocess_text(text2)

                ratio = fuzz.ratio(processed_text1, processed_text2) / 100
                partial_ratio = fuzz.partial_ratio(processed_text1, processed_text2) / 100
                token_sort_ratio = fuzz.token_sort_ratio(processed_text1, processed_text2) / 100
                token_set_ratio = fuzz.token_set_ratio(processed_text1, processed_text2) / 100

                type1 = self.get_product_type(text1)
                type2 = self.get_product_type(text2)

                type_multiplier = 1.2 if type1 == type2 and type1 != "other" else 1.0

                base_similarity = max(
                    ratio * 0.2 + token_sort_ratio * 0.4 + token_set_ratio * 0.4,
                    partial_ratio * 0.3 + token_sort_ratio * 0.7
                )
                
                return base_similarity * type_multiplier
            except Exception as e:
                logger.error(f"Ошибка при расчете сходства: {e}")
                return fuzz.ratio(text1.lower(), text2.lower()) / 100
        return await self.similarity_cache.get_or_create(cache_key, _calculate)

    async def _search_product_async(self, product_name: str) -> List[Tuple[str, str, float, float, str]]:
        """
            Асинхронный поиск товара во всех таблицах базы данных.
        """
        results = []
        tables = self.data_manager.get_all_table_names()
        
        tasks = []
        for table in tables:
            tasks.append(self._search_in_table_async(table, product_name))
        
        table_results = await asyncio.gather(*tasks)
        
        for table_result in table_results:
            results.extend(table_result)
        
        return sorted(results, key=lambda x: x[3], reverse=True)

    async def _search_in_table_async(self, table: str, product_name: str) -> List[Tuple[str, str, float, float, str]]:
        """
            Асинхронный поиск товара в конкретной таблице
        """
        results = []
        df = await asyncio.to_thread(self.data_manager.get_table_data, table)
        
        processed_product_name = await self.preprocess_text(product_name)
        product_type = self.get_product_type(product_name)
        
        processed_names = {}
        for idx, row in df.iterrows():
            name = row['Наименование']
            processed_names[idx] = await self.preprocess_text(name)
        
        exact_matches = []
        for idx, proc_name in processed_names.items():
            if proc_name == processed_product_name:
                exact_matches.append(idx)
                
        if exact_matches:
            for idx in exact_matches:
                product = df.iloc[idx]
                price = float(product.get('Цена с НДС', 0))
                description = str(product.get('Описание', ''))
                results.append((product['Наименование'], table, price, 1.0, description))
            return results
        distances, indices = await asyncio.to_thread(
            self.embedding_manager.search, table, "Наименование", product_name
        )

        if distances is not None and indices is not None:
            for dist, idx in zip(distances, indices):
                if idx < len(df):
                    product = df.iloc[idx]
                    found_name = product['Наименование']
                    fuzzy_similarity = await self.calculate_similarity(product_name, found_name)
                    if dist > 0.9:
                        final_similarity = max(dist, fuzzy_similarity)
                    else:
                        final_similarity = (dist * 0.3 + fuzzy_similarity * 0.7)

                    # Логируем результаты для отладки
                    logger.debug(f"""
                        Поиск для: {product_name} (тип: {product_type})
                        Обработанный запрос: {processed_product_name}
                        Найдено: {found_name}
                        Обработанная находка: {processed_names.get(idx, '')}
                        Эмбеддинг сходство: {dist:.3f}
                        Fuzzy similarity: {fuzzy_similarity:.3f}
                        Итоговое сходство: {final_similarity:.3f}
                    """)

                    price = float(product.get('Цена с НДС', 0))
                    description = str(product.get('Описание', ''))
                    results.append((found_name, table, price, float(final_similarity), description))

        results.sort(key=lambda x: x[3], reverse=True)
        
        filtered_results = [r for r in results if r[3] >= 0.5]

        high_similarity_results = [r for r in filtered_results if r[3] >= 0.8]
        if high_similarity_results:
            filtered_results = high_similarity_results[:3]
        else:
            filtered_results = filtered_results[:5] 
        logger.debug(f"Отфильтрованные результаты: {filtered_results}")
        return filtered_results

    async def validate_file_async(self, file_path: str) -> tuple[bool, str, Optional[pd.DataFrame]]:
        """
            Асинхронная валидация Excel файла.
    
        """
        try:
            excel_file = await asyncio.to_thread(pd.ExcelFile, file_path)
            all_sheets = []
            
            for sheet_name in excel_file.sheet_names:
                df = await asyncio.to_thread(pd.read_excel, file_path, sheet_name=sheet_name)
                df['sheet_name'] = sheet_name
                all_sheets.append(df)
            
            df = pd.concat(all_sheets, ignore_index=True)
            
            name_col = self._find_column(df, self.required_columns['name'])
            if not name_col:
                return False, "Не найдена колонка с наименованием", None
            
            quantity_col = self._find_column(df, self.required_columns['quantity'])
            
            if df[name_col].isna().all():
                return False, "Колонка с наименованием пуста", None
            
            return True, "", df
        except Exception as e:
            return False, f"Ошибка при чтении файла: {str(e)}", None

    async def _process_product_async(self, product_name: str, quantity: float) -> dict:
        """
            Асинхронная обработка одного продукта
        """
        try:
            search_results = await self._search_product_async(product_name)
            
            if search_results:
                best_match = search_results[0]
                found_name, table, price, similarity, description = best_match
                total_price = quantity * price
                
                return {
                    'Исходный товар': product_name,
                    'Найденный товар': found_name,
                    'Описание': description,
                    'Количество': quantity,
                    'Цена за штуку': price,
                    'Итоговая цена': total_price,
                    'Наша таблица': table,
                    'Сходство': similarity
                }
            else:
                return {
                    'Исходный товар': product_name,
                    'Найденный товар': 'Не найдено',
                    'Описание': '',
                    'Количество': quantity,
                    'Цена за штуку': 0,
                    'Итоговая цена': 0,
                    'Наша таблица': '',
                    'Сходство': 0
                }
        except Exception as e:
            logger.error(f"Error processing product {product_name}: {e}")
            return {
                'Исходный товар': product_name,
                'Найденный товар': 'Ошибка обработки',
                'Описание': str(e),
                'Количество': quantity,
                'Цена за штуку': 0,
                'Итоговая цена': 0,
                'Наша таблица': '',
                'Сходство': 0
            }

    async def process_file_async(self, input_file: str, output_file: str) -> tuple[bool, str]:
        """
            Асинхронная обработка входного Excel файла и создание выходного файла с результатами поиска.
        """
        try:
            await self._update_progress(0.1)
            is_valid, error_msg, df = await self.validate_file_async(input_file)
            if not is_valid:
                return False, error_msg

            name_col = self._find_column(df, self.required_columns['name'])
            quantity_col = self._find_column(df, self.required_columns['quantity'])

            result_data = []
            total_rows = len(df)
            processed_rows = 0

            for _, row in df.iterrows():
                product_name = row[name_col]
                
                if pd.isna(product_name) or str(product_name).strip() == '':
                    processed_rows += 1
                    continue
                
                quantity = 1.0
                if quantity_col:
                    try:
                        if not pd.isna(row[quantity_col]) and str(row[quantity_col]).strip() != '':
                            quantity = float(row[quantity_col])
                    except (ValueError, TypeError):
                        pass
                
                result = await self._process_product_async(str(product_name), quantity)
                result_data.append(result)
                processed_rows += 1
                
                progress = 0.1 + (0.7 * processed_rows / total_rows)
                await self._update_progress(progress)
            
            await self._update_progress(0.8)
            result_df = pd.DataFrame(result_data)
            
            def save_excel():
                result_df.to_excel(output_file, index=False)
            
            await asyncio.to_thread(save_excel)
            
            await self._update_progress(0.9)
            await asyncio.to_thread(self._format_excel, output_file)
            
            await self._update_progress(1.0)
            return True, "Файл успешно обработан"
        except Exception as e:
            logger.exception("Error in process_file_async")
            return False, f"Ошибка при обработке файла: {str(e)}"

    # Оставляем синхронные методы для обратной совместимости
    def _search_product(self, product_name: str) -> List[Tuple[str, str, float, float, str]]:
        """
            Синхронная версия поиска товара
        """
        return asyncio.run(self._search_product_async(product_name))

    def validate_file(self, file_path: str) -> tuple[bool, str, Optional[pd.DataFrame]]:
        """
            Синхронная версия валидации файла
        """
        return asyncio.run(self.validate_file_async(file_path))

    def process_file(self, input_file: str, output_file: str) -> tuple[bool, str]:
        """
            Синхронная версия обработки файла
        """
        return asyncio.run(self.process_file_async(input_file, output_file))

    def _format_excel(self, output_file: str):
        """
            Форматирование Excel файла с цветами и итоговой строкой
        """
        wb = load_workbook(output_file)
        ws = wb.active

        thin_border = Border(
            left   = Side(style='thin'),
            right  = Side(style='thin'),
            top    = Side(style='thin'),
            bottom = Side(style='thin')
        )

        green_fill  = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        red_fill    = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

        total_price_col = None
        description_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Итоговая цена':
                total_price_col = col
            elif ws.cell(row=1, column=col).value == 'Описание':
                description_col = col

        if total_price_col is None:
            raise ValueError("Колонка 'Итоговая цена' не найдена")
        if description_col is None:
            raise ValueError("Колонка 'Описание' не найдена")

        # Устанавливаем фиксированную ширину для всех колонок
        column_widths = {
            'Исходный товар': 30,
            'Найденный товар': 30,
            'Описание': 40,
            'Количество': 12,
            'Цена за штуку': 15,
            'Итоговая цена': 15,
            'Наша таблица': 20,
            'Сходство': 10
        }

        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=1, column=col).value
            if col_name in column_widths:
                ws.column_dimensions[get_column_letter(col)].width = column_widths[col_name]

        sum_formula_parts = []
        for row in range(1, ws.max_row + 1):  
            similarity = None
            if row > 1:
                similarity = float(ws.cell(row=row, column=ws.max_column).value)
            
            if row == 1:
                fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                font = Font(bold=True)
            elif similarity >= 1.0:
                fill = green_fill
                cell_ref = f"{get_column_letter(total_price_col)}{row}"
                sum_formula_parts.append(cell_ref)
            elif similarity >= 0.85:
                fill = yellow_fill
            else:
                fill = red_fill

            if row > 1:
                description_cell = ws.cell(row=row, column=description_col)
                if description_cell.value:
                    description = str(description_cell.value)
                    if len(description) > 300:
                        description = description[:300] + "..."
                    description_cell.value = description
                    description_cell.alignment = Alignment(wrap_text=True)

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.fill = fill
                if row == 1:  
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center')
                else: 
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="Итого:")
        
        if sum_formula_parts:
            sum_formula = f"=SUM({','.join(sum_formula_parts)})"
            ws.cell(row=total_row, column=total_price_col, value=sum_formula)
        
        bold_font = Font(bold=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.font = bold_font
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.border = thin_border

        wb.save(output_file)

    async def process_dataframe_async(self, df: pd.DataFrame, output_file: str) -> tuple[bool, str]:
        """
            Асинхронная обработка DataFrame и создание выходного файла
        """
        try:
            await self._update_progress(0.1)
            
            # Проверяем наличие необходимых колонок
            required_columns = ['Наименование', 'Цена с НДС', 'Описание']
            for col in required_columns:
                if col not in df.columns:
                    return False, f"Отсутствует обязательная колонка: {col}"
            
            # Сохраняем DataFrame в Excel
            def save_excel():
                # Используем стандартный метод to_excel
                df.to_excel(output_file, index=False)
                
                # Базовое форматирование с помощью openpyxl
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment, Border, Side
                
                wb = load_workbook(output_file)
                ws = wb.active
                
                # Форматируем заголовки
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Форматируем ячейки
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                
                # Автоматически подгоняем ширину колонок
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width
                
                wb.save(output_file)
            
            await asyncio.to_thread(save_excel)
            
            await self._update_progress(1.0)
            return True, "Файл успешно создан"
        except Exception as e:
            logger.exception("Error in process_dataframe_async")
            return False, f"Ошибка при создании файла: {str(e)}"

    async def create_pdf_from_dataframe(self, df: pd.DataFrame, output_file: str) -> tuple[bool, str]:
        """
            Создание PDF файла из DataFrame
        """
        try:
            # Функция для очистки и проверки данных
            def clean_and_validate_data(value):
                if pd.isna(value):
                    return ""
                
                # Преобразуем в строку
                value = str(value)
                
                # Убираем непечатаемые символы
                value = ''.join(char for char in value if char.isprintable())
                
                # Убираем лишние пробелы
                value = " ".join(value.strip().split())
                
                # Проверяем, что строка не состоит только из повторяющихся символов
                if len(set(value)) == 1 and len(value) > 3:
                    return ""
                    
                return value
            
            # Применяем очистку ко всем данным
            df = df.map(clean_and_validate_data)
            
            # Удаляем пустые строки
            df = df.replace("", pd.NA).dropna(how='all')
            
            if df.empty:
                return False, "Нет данных для создания PDF"
            
            # Создаем PDF документ
            pdf = SimpleDocTemplate(
                output_file,
                pagesize=A4,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Настройка стилей для русского текста
            styles = getSampleStyleSheet()
            styles['Normal'].fontName = 'Arial'  # Используем зарегистрированный шрифт
            styles['Normal'].encoding = 'UTF-8'
            
            # Подготавливаем данные для таблицы
            data = [df.columns.tolist()] + df.values.tolist()
            
            # Создаем таблицу
            table = Table(data)
            
            # Применяем стили к таблице
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),  # Используем Arial-Bold для заголовков
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            
            table.setStyle(style)
            
            # Добавляем таблицу в документ
            elements = [table]
            pdf.build(elements)
            
            return True, "PDF файл успешно создан"
        except Exception as e:
            logger.exception("Error in create_pdf_from_dataframe")
            return False, f"Ошибка при создании PDF файла: {str(e)}"