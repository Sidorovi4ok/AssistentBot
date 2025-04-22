import os
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from src.managers.manager_price import DataManager
from tqdm import tqdm

# --- –ù–ê–°–¢–†–û–ô–ö–ò ---
INPUT_EXCEL = 'tests/data/test_2.xlsx'
OUTPUT_EXCEL = 'sales_report.xlsx'
SIMILARITY_THRESHOLD = 80  # –ø–æ—Ä–æ–≥ —Å—Ö–æ–∂–µ—Å—Ç–∏ (0-100)
FUZZY_METHOD = fuzz.WRatio  # –∞–ª–≥–æ—Ä–∏—Ç–º fuzzywuzzy

# –í–æ–∑–º–æ–∂–Ω—ã–µ –Ω–∞–∑–≤–∞–Ω–∏—è –∫–æ–ª–æ–Ω–æ–∫
PRODUCT_COLUMNS = ['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞', '–¢–æ–≤–∞—Ä', '–ù–∞–∑–≤–∞–Ω–∏–µ', '–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ', 'Product', 'Item']
QUANTITY_COLUMNS = ['–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ', '–ö–æ–ª-–≤–æ', 'Quantity', '–ö–æ–ª', 'Qty']

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è DataManager
dm = DataManager.initialize('price-list.xlsx', update_db=False)


def preprocess_string(s: str) -> str:
    return str(s).strip().lower()


def find_product_in_db(product_name: str):
    exact_matches = []
    partial_matches = []
    clean_name = preprocess_string(product_name)

    for table in dm.get_all_table_names():
        df = dm.get_table_data(table)
        if '–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ' not in df.columns:
            continue

        for _, row in df.iterrows():
            db_name = preprocess_string(row['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ'])
            score = FUZZY_METHOD(clean_name, db_name)
            if score == 100:
                exact_matches.append(row)
            elif score >= SIMILARITY_THRESHOLD:
                partial_matches.append({'row': row, 'score': score})

    partial_matches.sort(key=lambda x: x['score'], reverse=True)
    return exact_matches, [pm['row'] for pm in partial_matches]


def process_input_sheets(input_file):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—Å–µ—Ö –ª–∏—Å—Ç–æ–≤ –≤—Ö–æ–¥–Ω–æ–≥–æ —Ñ–∞–π–ª–∞"""
    xls = pd.ExcelFile(input_file)
    all_data = []

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)

        # –ü–æ–∏—Å–∫ –∫–æ–ª–æ–Ω–∫–∏ —Å –Ω–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ–º —Ç–æ–≤–∞—Ä–∞
        product_col = next((col for col in PRODUCT_COLUMNS if col in df.columns), None)
        if not product_col:
            raise ValueError(f"–õ–∏—Å—Ç '{sheet_name}' –Ω–µ —Å–æ–¥–µ—Ä–∂–∏—Ç –ø–æ–¥—Ö–æ–¥—è—â–µ–π –∫–æ–ª–æ–Ω–∫–∏ —Å —Ç–æ–≤–∞—Ä–∞–º–∏")

        # –ü–µ—Ä–µ–∏–º–µ–Ω–æ–≤—ã–≤–∞–µ–º –≤ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ
        df.rename(columns={product_col: '–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞'}, inplace=True)

        # –ü–æ–∏—Å–∫ –∫–æ–ª–æ–Ω–∫–∏ —Å –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ–º
        quantity_col = next((col for col in QUANTITY_COLUMNS if col in df.columns), None)
        if not quantity_col:
            df['–ö–æ–ª-–≤–æ'] = 1
        else:
            # –ü–µ—Ä–µ–∏–º–µ–Ω–æ–≤—ã–≤–∞–µ–º –≤ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ
            df.rename(columns={quantity_col: '–ö–æ–ª-–≤–æ'}, inplace=True)

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞
        df['–ö–æ–ª-–≤–æ'] = df['–ö–æ–ª-–≤–æ'].apply(lambda x:
                                          int(float(x)) if pd.notnull(x) and float(x) > 0 else 1
                                          )

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏—Å—Ö–æ–¥–Ω–æ–µ –∏–º—è –ª–∏—Å—Ç–∞
        df['–ò—Å—Ç–æ—á–Ω–∏–∫'] = sheet_name
        all_data.append(df)

    return pd.concat(all_data, ignore_index=True)


# --- –û–ë–†–ê–ë–û–¢–ö–ê –í–•–û–î–ù–´–• –î–ê–ù–ù–´–• ---
combined_df = process_input_sheets(INPUT_EXCEL)
report_data = []

for _, row in tqdm(combined_df.iterrows(), total=len(combined_df), desc="üîç –û–±—Ä–∞–±–æ—Ç–∫–∞ —Ç–æ–≤–∞—Ä–æ–≤"):
    prod = row['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞']

    # –ü—Ä–æ–ø—É—Å–∫ –ø—É—Å—Ç—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
    if pd.isna(prod) or str(prod).strip() == '':
        continue
    prod = row['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞']
    quantity = row['–ö–æ–ª-–≤–æ']
    exact, partial = find_product_in_db(prod)

    base = {
        '–ò—Å—Ç–æ—á–Ω–∏–∫': row.get('–ò—Å—Ç–æ—á–Ω–∏–∫', ''),
        '–í–∞—à–∏ –ø—Ä–µ–¥–ª–æ–∂–µ–Ω–∏—è': prod,
        '–ù–∞—à–∏ —Ç–æ–≤–∞—Ä—ã': '',
        '–ö–æ–ª-–≤–æ': quantity,
        '–û–ø–∏—Å–∞–Ω–∏–µ': '',
        '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É': 0,
        '–ö–æ—ç—Ñ.': 0.0,
        '–°—Ç–∞—Ç—É—Å': 'red'
    }

    if exact:
        match = exact[0]
        base.update({
            '–ù–∞—à–∏ —Ç–æ–≤–∞—Ä—ã': match['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ'],
            '–û–ø–∏—Å–∞–Ω–∏–µ': match.get('–û–ø–∏—Å–∞–Ω–∏–µ', ''),
            '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É': match.get('–¶–µ–Ω–∞ —Å –ù–î–°', 0),
            '–ö–æ—ç—Ñ.': 1.0,
            '–°—Ç–∞—Ç—É—Å': 'green'
        })
    elif partial:
        match = partial[0]
        base.update({
            '–ù–∞—à–∏ —Ç–æ–≤–∞—Ä—ã': match['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ'],
            '–û–ø–∏—Å–∞–Ω–∏–µ': match.get('–û–ø–∏—Å–∞–Ω–∏–µ', ''),
            '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É': match.get('–¶–µ–Ω–∞ —Å –ù–î–°', 0),
            '–°—Ç–∞—Ç—É—Å': 'yellow'
        })

    report_data.append(base)

# --- –§–û–†–ú–ò–†–û–í–ê–ù–ò–ï –û–¢–ß–Å–¢–ê ---
report_df = pd.DataFrame(report_data)
report_df.insert(0, '‚Ññ', range(1, len(report_df) + 1))
report_df['–ò—Ç–æ–≥–æ–≤–∞—è —Ü–µ–Ω–∞'] = (
        report_df['–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É'] *
        report_df['–ö–æ–ª-–≤–æ'] *
        report_df['–ö–æ—ç—Ñ.']
)

report_df.rename(columns={'–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É': '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É, ‚ÇΩ'}, inplace=True)
export_columns = ['‚Ññ', '–ò—Å—Ç–æ—á–Ω–∏–∫', '–í–∞—à–∏ –ø—Ä–µ–¥–ª–æ–∂–µ–Ω–∏—è', '–ù–∞—à–∏ —Ç–æ–≤–∞—Ä—ã',
                  '–ö–æ–ª-–≤–æ', '–û–ø–∏—Å–∞–Ω–∏–µ', '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É, ‚ÇΩ',
                  '–ö–æ—ç—Ñ.', '–ò—Ç–æ–≥–æ–≤–∞—è —Ü–µ–Ω–∞']

# --- –ó–ê–ü–ò–°–¨ –í EXCEL ---
wb = Workbook()
ws = wb.active
ws.append(export_columns)

# –°—Ç–∏–ª–∏ –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—è
fills = {
    'green': PatternFill(start_color='C6EFCE', fill_type='solid'),
    'yellow': PatternFill(start_color='FFEB9C', fill_type='solid'),
    'red': PatternFill(start_color='FFC7CE', fill_type='solid'),
}
bold_font = Font(bold=True)
border = Border(top=Side(style='medium'), bottom=Side(style='medium'))

for idx, row in report_df.iterrows():
    ws.append([row[col] for col in export_columns])
    for cell in ws[ws.max_row]:
        cell.fill = fills[row['–°—Ç–∞—Ç—É—Å']]

# –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç—Ä–æ–∫–∞
total = report_df.loc[report_df['–°—Ç–∞—Ç—É—Å'] == 'green', '–ò—Ç–æ–≥–æ–≤–∞—è —Ü–µ–Ω–∞'].sum()
ws.append(['–ò—Ç–æ–≥–æ'] + [''] * (len(export_columns) - 2) + [total])

# –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∏—Ç–æ–≥–æ–≤–æ–π —Å—Ç—Ä–æ–∫–∏
for cell in ws[ws.max_row]:
    cell.font = bold_font
    cell.border = border

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ —à–∏—Ä–∏–Ω—ã —Å—Ç–æ–ª–±—Ü–æ–≤
col_widths = {
    'A': 8, 'B': 20, 'C': 40, 'D': 40,
    'E': 10, 'F': 80, 'G': 18, 'H': 10, 'I': 15
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUTPUT_EXCEL)
print(f'–û—Ç—á–µ—Ç —É—Å–ø–µ—à–Ω–æ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω: {OUTPUT_EXCEL}')